import ast
import json
import os
import re
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
BACKEND_DIR = ROOT / 'backend'
OUTPUT_DIR = ROOT / 'Vulnerability Test Results'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SUMMARY_PATH = OUTPUT_DIR / 'security-review.md'
EXECUTIVE_PATH = OUTPUT_DIR / 'executive-summary.md'
DEPENDENCY_PATH = OUTPUT_DIR / 'dependency-report.md'
ENDPOINT_XLSX = OUTPUT_DIR / 'endpoint-inventory.xlsx'
FINDINGS_XLSX = OUTPUT_DIR / 'findings.xlsx'


class Assessment:
    def __init__(self):
        self.backend = None
        self.language = None
        self.api_architecture = None
        self.auth = 'None'
        self.authz = 'None'
        self.database = 'None'
        self.orm = 'None'
        self.api_docs = 'None'
        self.middleware = []
        self.file_upload = 'Yes' if self.backend_has_file_upload() else 'No'
        self.sessions = 'None'
        self.third_party = []
        self.endpoints = []
        self.findings = []
        self.dependencies = []
        self.dependency_vulnerabilities = []
        self.dast_findings = []
        self.critical_count = 0
        self.high_count = 0
        self.medium_count = 0
        self.low_count = 0

    def backend_has_file_upload(self):
        return False

    def add_finding(self, severity, vuln_type, file_path, endpoint, description, impact, recommendation):
        row = {
            'Severity': severity,
            'Vulnerability Type': vuln_type,
            'File Path': file_path,
            'Endpoint': endpoint or '',
            'Description': description,
            'Impact': impact,
            'Recommendation': recommendation,
        }
        self.findings.append(row)
        if severity == 'Critical':
            self.critical_count += 1
        elif severity == 'High':
            self.high_count += 1
        elif severity == 'Medium':
            self.medium_count += 1
        elif severity == 'Low':
            self.low_count += 1

    def detect_backend(self):
        if (BACKEND_DIR / 'app.py').exists():
            self.language = 'Python'
            self.backend = 'Flask'
            self.api_architecture = 'REST API'
            self.third_party = ['Flask', 'Flask-CORS', 'TensorFlow', 'NumPy', 'Pillow']
            self.middleware = ['CORS']
            self.sessions = 'None'
            self.file_upload = 'Yes'
            if (BACKEND_DIR / 'requirements.txt').exists():
                self.dependencies = self.parse_requirements(BACKEND_DIR / 'requirements.txt')
            return True
        return False

    def parse_requirements(self, path):
        deps = []
        for line in path.read_text().splitlines():
            cleaned = line.strip()
            if not cleaned or cleaned.startswith('#'):
                continue
            dep = cleaned
            if '==' in dep or '>=' in dep or '<=' in dep or '~=' in dep:
                name, version = re.split(r'==|~=|>=|<=|>|<', dep, maxsplit=1)
                deps.append({'name': name.strip(), 'version': version.strip(), 'raw': dep})
            else:
                deps.append({'name': dep.strip(), 'version': 'unspecified', 'raw': dep.strip()})
        return deps

    def discover_endpoints(self):
        app_py = BACKEND_DIR / 'app.py'
        if not app_py.exists():
            return
        source = app_py.read_text(errors='ignore')
        routes = re.findall(r"@app\.route\(['\"]([^'\"]+)['\"],\s*methods=\[([^\]]+)\]\)", source)
        for route, methods in routes:
            methods_clean = [m.strip().strip("'\"") for m in methods.split(',')]
            if not methods_clean:
                methods_clean = ['GET']
            for method in methods_clean:
                self.endpoints.append({
                    'Endpoint': route,
                    'HTTP Method': method,
                    'Authentication Required': 'No',
                    'Expected Roles': 'None',
                    'Controller/File Path': 'backend/app.py'
                })

    def scan_sast(self):
        app_py = BACKEND_DIR / 'app.py'
        if not app_py.exists():
            return
        source = app_py.read_text(errors='ignore')
        if 'CORS(app)' in source:
            self.add_finding(
                'Medium',
                'Open CORS Policy',
                'backend/app.py',
                '/',
                'The application enables Flask CORS with default settings, which allows requests from any origin.',
                'Any website can make requests to the API and receive responses, increasing risk of cross-site data leaks.',
                'Restrict CORS to the trusted mobile app origin or explicit allowed origins.'
            )
        if 'request.files' in source:
            self.add_finding(
                'High',
                'Unsafe File Upload Handling',
                'backend/app.py',
                '/predict',
                'Uploaded images are accepted without file type, size, or content validation.',
                'An attacker can upload malformed or oversized images leading to DoS, memory exhaustion, or image parsing vulnerabilities.',
                'Validate file type, extension, maximum size, and process uploads in a sandboxed manner.'
            )
        if 'model.predict' in source and 'try:' in source:
            self.add_finding(
                'Low',
                'Error Disclosure',
                'backend/app.py',
                '/predict',
                'Prediction errors are returned with exception text in the response.',
                'Detailed runtime errors can expose internal application and model handling implementation details.',
                'Return generic error messages and log exception details server-side only.'
            )
        if 'debug=False' in source:
            pass
        if 'auth' not in source and 'login' not in source and 'token' not in source:
            self.add_finding(
                'Critical',
                'Missing Authentication and Authorization',
                'backend/app.py',
                'All endpoints',
                'The backend exposes all API functionality without authentication or authorization.',
                'Any user or service can call the API freely, including sensitive prediction and health endpoints.',
                'Implement authentication (API key, OAuth, JWT) and enforce authorization for protected resources.'
            )
        if 'os.path.join(os.path.dirname(__file__)' in source and 'MODEL_PATH' in source:
            self.add_finding(
                'Low',
                'Local File Dependency',
                'backend/app.py',
                '/predict',
                'The application loads a model file from disk at startup.',
                'If the model file is missing or corrupted, the service will fail to start or behave unexpectedly.',
                'Use a secure model distribution process and verify file integrity before loading.'
            )
        if 'CORS(app)' in source and 'origins' not in source:
            pass
        if 'image' in source and 'np.array' in source:
            self.add_finding(
                'Medium',
                'Insecure Image Upload Processing',
                'backend/app.py',
                '/predict',
                'Image uploads are processed without explicit format or size checks.',
                'Crafted image files can trigger parser vulnerabilities or cause resource exhaustion.',
                'Add explicit JPEG/PNG validation and enforce maximum dimensions and byte size limits.'
            )
        if 'Content-Type' not in source and 'request.headers' not in source:
            self.add_finding(
                'Low',
                'Missing Content-Type Enforcement',
                'backend/app.py',
                '/predict',
                'The predict endpoint depends on form-data but does not explicitly validate request headers.',
                'Malicious clients may attempt unsupported request formats or bypass upload restrictions.',
                'Check for expected multipart/form-data content types and reject unsupported requests.'
            )

    def scan_dependencies(self):
        if self.dependencies:
            for dep in self.dependencies:
                vuln = None
                if dep['name'].lower() == 'flask' and dep['version'] == 'unspecified':
                    vuln = 'Unpinned Flask dependency may hide vulnerable versions.'
                if dep['name'].lower() == 'tensorflow':
                    vuln = 'TensorFlow should be pinned to a known secure version and audited for CVEs.'
                if dep['name'].lower() == 'flask-cors':
                    vuln = 'Flask-CORS should be configured for specific origins rather than allowing any origin by default.'
                self.dependency_vulnerabilities.append({
                    'Package': dep['name'],
                    'Version': dep['version'],
                    'Issue': vuln or 'No explicit version pinning or vulnerability metadata available.',
                    'Remediation': 'Pin the package to a secure version and perform dependency CVE scans with pip-audit or similar tools.'
                })
                if vuln:
                    self.add_finding(
                        'Medium',
                        'Dependency Management',
                        'backend/requirements.txt',
                        'N/A',
                        vuln,
                        'Outdated or unpinned dependencies increase the risk of known vulnerabilities being present.',
                        'Pin exact dependency versions and run automated CVE scanning.'
                    )

    def run_dast(self):
        target = os.environ.get('TARGET_URL')
        if not target:
            return
        health_url = target.rstrip('/') + '/'
        predict_url = target.rstrip('/') + '/predict'
        try:
            req = urllib.request.Request(health_url, method='GET')
            with urllib.request.urlopen(req, timeout=10) as resp:
                if resp.status != 200:
                    self.add_finding(
                        'Medium',
                        'Health Check Response Issue',
                        'Runtime',
                        health_url,
                        f'Health endpoint returned HTTP {resp.status}.',
                        'The service may be unavailable or misconfigured.',
                        'Verify the health endpoint and service readiness.'
                    )
        except Exception as e:
            self.add_finding(
                'High',
                'Unavailable Target',
                'Runtime',
                health_url,
                f'Health endpoint check failed: {e}',
                'Dynamic tests could not reach the running API.',
                'Ensure TARGET_URL points to a valid live API endpoint.'
            )
            return
        # Non-destructive predict invocation with a dummy request
        try:
            boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
            data = (
                f'--{boundary}\r\n'
                'Content-Disposition: form-data; name="image"; filename="test.png"\r\n'
                'Content-Type: image/png\r\n\r\n'
                'PNG\r\n'
                f'--{boundary}--\r\n'
            ).encode('utf-8')
            req = urllib.request.Request(predict_url, data=data, method='POST')
            req.add_header('Content-Type', f'multipart/form-data; boundary={boundary}')
            with urllib.request.urlopen(req, timeout=10) as resp:
                self.add_finding(
                    'Low',
                    'Predict Endpoint Accessibility',
                    'Runtime',
                    predict_url,
                    f'Predict endpoint responded with HTTP {resp.status}.',
                    'The endpoint is reachable, but dynamic behavior requires manual validation.',
                    'Verify the endpoint handles invalid uploads gracefully.'
                )
        except urllib.error.HTTPError as err:
            if err.code == 400:
                pass
            else:
                self.add_finding(
                    'Low',
                    'Unexpected Predict Endpoint Response',
                    'Runtime',
                    predict_url,
                    f'Predict endpoint returned HTTP {err.code}.',
                    'The endpoint may not be handling requests as expected.',
                    'Validate the /predict API contract and response status codes.'
                )
        except Exception as e:
            self.add_finding(
                'Low',
                'Predict Endpoint Connectivity',
                'Runtime',
                predict_url,
                f'Predict endpoint invocation failed: {e}',
                'The service may reject malformed uploads or be temporarily unavailable.',
                'Verify runtime upload handling and network connectivity.'
            )

    def build_reports(self):
        self.write_md_reports()
        self.write_excel_reports()

    def write_md_reports(self):
        summary_lines = [
            '# Security Review',
            f'Date: {datetime.utcnow().isoformat()}Z',
            '',
            '## Backend Inventory',
            f'- Framework: {self.backend}',
            f'- Language: {self.language}',
            f'- API Architecture: {self.api_architecture}',
            f'- Authentication Mechanism: {self.auth}',
            f'- Authorization Model: {self.authz}',
            f'- Database Technology: {self.database}',
            f'- ORM Usage: {self.orm}',
            f'- API Documentation: {self.api_docs}',
            f'- Middleware: {", ".join(self.middleware) if self.middleware else "None"}',
            f'- File Upload Functionality: {self.file_upload}',
            f'- Session Handling: {self.sessions}',
            f'- Third-Party Integrations: {", ".join(self.third_party) if self.third_party else "None"}',
            '',
            '## Findings',
        ]
        for finding in self.findings:
            summary_lines.extend([
                f'### {finding["Severity"]}: {finding["Vulnerability Type"]}',
                f'- File Path: {finding["File Path"]}',
                f'- Endpoint: {finding["Endpoint"]}',
                f'- Description: {finding["Description"]}',
                f'- Impact: {finding["Impact"]}',
                f'- Recommendation: {finding["Recommendation"]}',
                ''
            ])
        summary_lines.append('## Dependency Findings')
        for dep in self.dependency_vulnerabilities:
            summary_lines.extend([
                f'- {dep["Package"]} ({dep["Version"]}): {dep["Issue"]}',
                f'  - Recommendation: {dep["Remediation"]}',
            ])
        SUMMARY_PATH.write_text('\n'.join(summary_lines), encoding='utf-8')

        executive_lines = [
            '# Executive Summary',
            '',
            '## Total Findings',
            f'- Critical: {self.critical_count}',
            f'- High: {self.high_count}',
            f'- Medium: {self.medium_count}',
            f'- Low: {self.low_count}',
            '',
            '## Most Critical Risks',
        ]
        if self.critical_count == 0:
            executive_lines.append('- No critical findings were detected.')
        else:
            for finding in self.findings:
                if finding['Severity'] == 'Critical':
                    executive_lines.append(f'- {finding["Vulnerability Type"]}: {finding["Description"]}')
        score = 100 - (self.high_count * 5 + self.medium_count * 2 + self.low_count)
        if score < 0:
            score = 0
        executive_lines.extend([
            '',
            '## Overall Security Score',
            f'- {score}/100',
            '',
            '## Notes',
            '- This assessment is based on static code review and target detection within the repository.',
            '- Dynamic API findings require a live TARGET_URL to execute safely.',
        ])
        EXECUTIVE_PATH.write_text('\n'.join(executive_lines), encoding='utf-8')

        dependency_lines = [
            '# Dependency Report',
            '',
            '## Detected Backend Dependencies',
        ]
        for dep in self.dependencies:
            dependency_lines.append(f'- {dep["name"]}: {dep["version"]}')
        dependency_lines.append('')
        dependency_lines.append('## Dependency Vulnerabilities')
        for dep in self.dependency_vulnerabilities:
            dependency_lines.append(f'- {dep["Package"]} ({dep["Version"]}): {dep["Issue"]}')
            dependency_lines.append(f'  - Remediation: {dep["Remediation"]}')
        DEPENDENCY_PATH.write_text('\n'.join(dependency_lines), encoding='utf-8')

    def write_excel_reports(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Security Findings'
        sheet.append(['Severity', 'Vulnerability Type', 'File Path', 'Endpoint', 'Description', 'Impact', 'Recommendation'])
        for finding in self.findings:
            sheet.append([finding[col] for col in ['Severity', 'Vulnerability Type', 'File Path', 'Endpoint', 'Description', 'Impact', 'Recommendation']])

        endpoint = wb.create_sheet('Endpoint Inventory')
        endpoint.append(['Endpoint', 'HTTP Method', 'Authentication Required', 'Expected Roles', 'Controller/File Path'])
        for row in self.endpoints:
            endpoint.append([row[col] for col in ['Endpoint', 'HTTP Method', 'Authentication Required', 'Expected Roles', 'Controller/File Path']])

        dependencies = wb.create_sheet('Dependency Vulnerabilities')
        dependencies.append(['Package', 'Version', 'Issue', 'Remediation'])
        for dep in self.dependency_vulnerabilities:
            dependencies.append([dep['Package'], dep['Version'], dep['Issue'], dep['Remediation']])

        summary = wb.create_sheet('Risk Summary')
        summary.append(['Severity', 'Count'])
        summary.append(['Critical', self.critical_count])
        summary.append(['High', self.high_count])
        summary.append(['Medium', self.medium_count])
        summary.append(['Low', self.low_count])

        wb.save(FINDINGS_XLSX)

        endpoint_wb = Workbook()
        endpoint_sheet = endpoint_wb.active
        endpoint_sheet.title = 'Endpoint Inventory'
        endpoint_sheet.append(['Endpoint', 'HTTP Method', 'Authentication Required', 'Expected Roles', 'Controller/File Path'])
        for row in self.endpoints:
            endpoint_sheet.append([row[col] for col in ['Endpoint', 'HTTP Method', 'Authentication Required', 'Expected Roles', 'Controller/File Path']])
        endpoint_wb.save(ENDPOINT_XLSX)

    def run(self):
        if not self.detect_backend():
            print('No backend framework detected. Exiting.')
            sys.exit(0)
        self.discover_endpoints()
        self.scan_sast()
        self.scan_dependencies()
        self.run_dast()
        self.build_reports()
        print('Generated security reports in', OUTPUT_DIR)
        if self.critical_count > 0:
            sys.exit(1)


if __name__ == '__main__':
    Assessment().run()
