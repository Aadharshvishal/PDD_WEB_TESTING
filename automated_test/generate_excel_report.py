"""
generate_excel_report.py
Reads automated_test/report.json and writes a styled Excel workbook.
Requires: openpyxl  (pip install openpyxl)
"""
import json, sys, pathlib, subprocess
from datetime import datetime

# Ensure openpyxl is installed before importing
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

REPORT_JSON = pathlib.Path(__file__).parent / "report.json"
OUT_XLSX    = pathlib.Path(__file__).parent / "dast_report.xlsx"

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "critical":      "C0392B",   # deep red
    "high":          "E67E22",   # orange
    "medium":        "F1C40F",   # yellow
    "low":           "2980B9",   # blue
    "info":          "95A5A6",   # grey
    "ok":            "27AE60",   # green
    "header_bg":     "1A1A2E",   # dark navy
    "header_fg":     "FFFFFF",
    "subheader_bg":  "16213E",
    "subheader_fg":  "E0E0E0",
    "alt_row":       "F5F7FA",
    "white":         "FFFFFF",
    "title_bg":      "0F3460",
    "title_fg":      "E94560",
    "accent":        "E94560",
}

SEV_LABEL = {
    "critical": "🔴 CRITICAL",
    "high":     "🟠 HIGH",
    "medium":   "🟡 MEDIUM",
    "low":      "🔵 LOW",
    "info":     "⚪ INFO",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def apply_header(ws, row, col, value, bg=C["header_bg"], fg=C["header_fg"],
                 bold=True, size=11, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = fill(bg)
    cell.font      = font(bold=bold, color=fg, size=size)
    cell.alignment = center() if align == "center" else left()
    cell.border    = thick_border()
    return cell

# ── Load data ─────────────────────────────────────────────────────────────────
data = json.loads(REPORT_JSON.read_text(encoding="utf-8-sig"))
findings_only = [r for r in data if r["finding"]]

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Executive Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["F"].width = 3

# Title block
ws1.row_dimensions[1].height = 10
ws1.row_dimensions[2].height = 42
ws1.row_dimensions[3].height = 24
ws1.row_dimensions[4].height = 10

ws1.merge_cells("B2:E2")
title_cell = ws1["B2"]
title_cell.value     = "DAST Security Audit Report"
title_cell.fill      = fill(C["title_bg"])
title_cell.font      = Font(bold=True, color=C["title_fg"], size=22, name="Calibri")
title_cell.alignment = center()
title_cell.border    = thick_border()

ws1.merge_cells("B3:E3")
sub_cell = ws1["B3"]
sub_cell.value     = f"OSCC Flask API  ·  http://localhost:5000  ·  {datetime.now().strftime('%Y-%m-%d')}"
sub_cell.fill      = fill(C["subheader_bg"])
sub_cell.font      = font(bold=False, color=C["subheader_fg"], size=11, italic=True)
sub_cell.alignment = center()
sub_cell.border    = thick_border()

# Stats cards
stats = [
    ("Endpoints", "2"),
    ("Tests Run",  str(len(data))),
    ("Findings",   str(len(findings_only))),
]
sev_counts = {}
for r in data:
    if r["finding"]:
        sev_counts[r["severity"]] = sev_counts.get(r["severity"], 0) + 1

sev_order = ["critical", "high", "medium", "low", "info"]
card_colors = {
    "critical": C["critical"],
    "high":     C["high"],
    "medium":   C["medium"],
    "low":      C["low"],
    "info":     C["info"],
}

row = 5
ws1.row_dimensions[row].height = 14

# Stats header
for col, (label, val) in zip([2, 3, 4], stats):
    ws1.row_dimensions[row+1].height = 40
    ws1.row_dimensions[row+2].height = 24
    c1 = ws1.cell(row=row+1, column=col, value=val)
    c1.fill = fill(C["title_bg"])
    c1.font = Font(bold=True, color=C["title_fg"], size=26, name="Calibri")
    c1.alignment = center()
    c1.border = thick_border()
    c2 = ws1.cell(row=row+2, column=col, value=label)
    c2.fill = fill(C["subheader_bg"])
    c2.font = font(bold=True, color=C["subheader_fg"], size=10)
    c2.alignment = center()
    c2.border = thick_border()

row = 9
ws1.row_dimensions[row].height = 14

# Severity breakdown
for col_offset, sev in enumerate(sev_order, start=0):
    col = 2 + col_offset % 4
    ws1.row_dimensions[row + col_offset // 4 * 3 + 1].height = 36
    ws1.row_dimensions[row + col_offset // 4 * 3 + 2].height = 20
    r1 = row + (col_offset // 4) * 3 + 1
    r2 = r1 + 1
    count = sev_counts.get(sev, 0)
    c1 = ws1.cell(row=r1, column=col, value=count)
    c1.fill = fill(card_colors[sev])
    c1.font = Font(bold=True, color="FFFFFF", size=24, name="Calibri")
    c1.alignment = center()
    c1.border = thick_border()
    c2 = ws1.cell(row=r2, column=col, value=sev.upper())
    c2.fill = fill(card_colors[sev])
    c2.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    c2.alignment = center()
    c2.border = thick_border()

# Scope table
scope_row = 18
ws1.row_dimensions[scope_row].height = 14
apply_header(ws1, scope_row+1, 2, "Scope & Configuration", bg=C["header_bg"],
             fg=C["header_fg"], size=12, align="left")
ws1.merge_cells(f"B{scope_row+1}:E{scope_row+1}")

scope_data = [
    ("Target URL",         "http://localhost:5000"),
    ("Codebase",           "d:/d/PDD_Web/backend/app.py"),
    ("Framework",          "Flask + TensorFlow (OSCC Medical AI)"),
    ("Auth Scheme Found",  "NONE — no authentication implemented"),
    ("OpenAPI Spec",       "Not served"),
    ("Endpoints Tested",   "GET /   ·   POST /predict"),
    ("Test Date",          datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Tester",             "Antigravity DAST Runner (Automated)"),
]

for i, (k, v) in enumerate(scope_data):
    r = scope_row + 2 + i
    ws1.row_dimensions[r].height = 20
    ck = ws1.cell(row=r, column=2, value=k)
    ck.fill = fill(C["subheader_bg"] if i % 2 == 0 else "1E2C4A")
    ck.font = font(bold=True, color=C["subheader_fg"], size=10)
    ck.alignment = left(wrap=False)
    ck.border = border()
    cv = ws1.cell(row=r, column=3, value=v)
    cv.fill = fill(C["alt_row"] if i % 2 == 0 else C["white"])
    cv.font = font(color="222222", size=10)
    cv.alignment = left(wrap=False)
    cv.border = border()
    ws1.merge_cells(f"C{r}:E{r}")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — All Test Results
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Test Results")
ws2.sheet_view.showGridLines = False

COL_WIDTHS = [4, 32, 8, 18, 12, 14, 12, 18, 12, 60, 4]
COL_NAMES  = ["#", "Endpoint", "Method", "Role", "Status",
              "Expected", "Finding", "Category", "Severity", "Note"]

for i, w in enumerate(COL_WIDTHS, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 10
ws2.row_dimensions[2].height = 38
ws2.merge_cells("A2:K2")
title = ws2["A2"]
title.value     = "All Test Results — Raw DAST Data"
title.fill      = fill(C["title_bg"])
title.font      = Font(bold=True, color=C["title_fg"], size=18, name="Calibri")
title.alignment = center()

ws2.row_dimensions[3].height = 28
for col, name in enumerate(COL_NAMES, 1):
    apply_header(ws2, 3, col, name, size=10)
apply_header(ws2, 3, 11, "", bg=C["header_bg"])  # spacer

for row_idx, rec in enumerate(data, start=4):
    r = row_idx
    ws2.row_dimensions[r].height = 22
    is_alt = row_idx % 2 == 0

    sev = rec.get("severity", "info")
    is_finding = rec.get("finding", False)

    row_bg = C["alt_row"] if is_alt else C["white"]

    values = [
        row_idx - 3,
        rec["endpoint"].replace("http://localhost:5000", ""),
        rec["method"],
        rec["role"],
        rec.get("status", "—"),
        str(rec.get("expected_status", "")),
        "YES" if is_finding else "no",
        rec.get("test_category", ""),
        sev.upper(),
        rec.get("note", ""),
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws2.cell(row=r, column=col_idx, value=val)
        cell.border    = border()
        cell.alignment = center() if col_idx in (1,3,5,6,7,9) else left(wrap=(col_idx == 10))

        if col_idx == 7 and is_finding:
            cell.fill = fill(C[sev] if sev in C else C["info"])
            cell.font = font(bold=True, color="FFFFFF", size=10)
        elif col_idx == 9:
            cell.fill = fill(C[sev] if sev in C else C["info"])
            cell.font = font(bold=True, color="FFFFFF", size=9)
        else:
            cell.fill = fill(row_bg)
            cell.font = font(size=9,
                             bold=(col_idx == 2),
                             color="111111")

    # Status code colouring
    status_cell = ws2.cell(row=r, column=5)
    st = rec.get("status")
    if st is not None:
        if st < 300:
            status_cell.font = font(bold=True, color="27AE60", size=9)
        elif st < 400:
            status_cell.font = font(bold=True, color="E67E22", size=9)
        elif st < 500:
            status_cell.font = font(bold=True, color="C0392B", size=9)

# Freeze top rows
ws2.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Findings Only
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Findings")
ws3.sheet_view.showGridLines = False

for i, w in enumerate(COL_WIDTHS, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.row_dimensions[1].height = 10
ws3.row_dimensions[2].height = 38
ws3.merge_cells("A2:K2")
t3 = ws3["A2"]
t3.value     = f"Security Findings ({len(findings_only)} total)"
t3.fill      = fill(C["critical"])
t3.font      = Font(bold=True, color="FFFFFF", size=18, name="Calibri")
t3.alignment = center()

ws3.row_dimensions[3].height = 28
for col, name in enumerate(COL_NAMES, 1):
    apply_header(ws3, 3, col, name, size=10)
apply_header(ws3, 3, 11, "", bg=C["header_bg"])

sorted_findings = sorted(findings_only,
    key=lambda x: sev_order.index(x.get("severity","info")) if x.get("severity","info") in sev_order else 99)

for row_idx, rec in enumerate(sorted_findings, start=4):
    r = row_idx
    ws3.row_dimensions[r].height = 28
    sev = rec.get("severity", "info")
    sev_bg = C.get(sev, C["info"])

    values = [
        row_idx - 3,
        rec["endpoint"].replace("http://localhost:5000", ""),
        rec["method"],
        rec["role"],
        rec.get("status", "—"),
        str(rec.get("expected_status", "")),
        "FINDING",
        rec.get("test_category", ""),
        sev.upper(),
        rec.get("note", ""),
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws3.cell(row=r, column=col_idx, value=val)
        cell.border    = border()

        if col_idx == 1:
            cell.fill = fill(sev_bg)
            cell.font = font(bold=True, color="FFFFFF", size=10)
            cell.alignment = center()
        elif col_idx == 9:
            cell.fill = fill(sev_bg)
            cell.font = font(bold=True, color="FFFFFF", size=9)
            cell.alignment = center()
        elif col_idx == 7:
            cell.fill = fill(sev_bg)
            cell.font = font(bold=True, color="FFFFFF", size=9)
            cell.alignment = center()
        elif col_idx == 10:
            cell.fill = fill("FFF9F9" if sev == "critical" else "FFFDF5" if sev == "high" else "FFFFF5")
            cell.font = font(size=9, color="111111")
            cell.alignment = left(wrap=True)
        else:
            cell.fill = fill("FFF9F9" if sev == "critical" else "FFFDF5" if sev == "high" else "FFFFF5")
            cell.font = font(size=9, color="111111",
                             bold=(col_idx in (2, 3)))
            cell.alignment = center() if col_idx in (3,5,6) else left()

ws3.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Remediation Checklist
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Remediation Checklist")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 4
ws4.column_dimensions["B"].width = 6
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 55
ws4.column_dimensions["F"].width = 22
ws4.column_dimensions["G"].width = 14
ws4.column_dimensions["H"].width = 4

ws4.row_dimensions[1].height = 10
ws4.row_dimensions[2].height = 38
ws4.merge_cells("B2:G2")
t4 = ws4["B2"]
t4.value     = "Remediation Checklist"
t4.fill      = fill(C["title_bg"])
t4.font      = Font(bold=True, color=C["title_fg"], size=18, name="Calibri")
t4.alignment = center()

ws4.row_dimensions[3].height = 26
for col, hdr in enumerate(["#", "Severity", "Category", "Description", "Fix / Action", "Status"], start=2):
    apply_header(ws4, 3, col, hdr, size=10)

REMEDIATIONS = [
    ("critical", "authn_bypass",
     "No authentication on POST /predict",
     "Add JWT middleware (@require_auth decorator). Set JWT_SECRET env var. Return 401 for missing/invalid tokens.",
     "🔲 TODO"),
    ("critical", "token_tampering",
     "Tampered JWTs accepted (alg:none, fake sig)",
     "Implement PyJWT with algorithm whitelist: jwt.decode(token, secret, algorithms=['HS256']). Reject alg:none.",
     "🔲 TODO"),
    ("high", "authz_privesc",
     "No RBAC — all roles get equal access",
     "Add @require_role('clinician','admin') decorator after auth is in place.",
     "🔲 TODO"),
    ("high", "hardcoded_creds",
     "Hardcoded LAN IP in mobile_app/config.js",
     "Add mobile_app/config.js to .gitignore. start_oscc.bat already auto-generates it — never commit.",
     "🔲 TODO"),
    ("medium", "rate_limiting",
     "No rate limiting on POST /predict",
     "Install Flask-Limiter. Apply @limiter.limit('10 per minute') to /predict to prevent inference-flood DoS.",
     "🔲 TODO"),
    ("medium", "misc",
     "4 security headers missing",
     "Add after_request hook: X-Content-Type-Options: nosniff, X-Frame-Options: DENY, CSP: default-src 'none', HSTS.",
     "🔲 TODO"),
    ("low", "misc",
     "raw_score (internal ML logit) exposed",
     "Remove raw_score key from the /predict JSON response. Expose only risk_level, confidence, recommendation, color_code.",
     "🔲 TODO"),
]

for row_idx, (sev, cat, desc, fix, status) in enumerate(REMEDIATIONS, start=4):
    r = row_idx
    ws4.row_dimensions[r].height = 52
    sev_bg = C.get(sev, C["info"])
    row_bg = "FFF9F9" if sev == "critical" else "FFFDF5" if sev == "high" else "FFFFF5" if sev == "medium" else "F5F9FF"

    ws4.cell(row=r, column=2, value=row_idx-3).fill = fill(sev_bg)
    ws4.cell(row=r, column=2).font = font(bold=True, color="FFFFFF", size=10)
    ws4.cell(row=r, column=2).alignment = center()
    ws4.cell(row=r, column=2).border = thick_border()

    ws4.cell(row=r, column=3, value=sev.upper()).fill = fill(sev_bg)
    ws4.cell(row=r, column=3).font = font(bold=True, color="FFFFFF", size=9)
    ws4.cell(row=r, column=3).alignment = center()
    ws4.cell(row=r, column=3).border = thick_border()

    for col, val in [(4, cat), (5, desc), (6, fix), (7, status)]:
        cell = ws4.cell(row=r, column=col, value=val)
        cell.fill      = fill(row_bg)
        cell.font      = font(size=9, color="111111", bold=(col == 5))
        cell.alignment = left(wrap=True)
        cell.border    = border()

ws4.freeze_panes = "B4"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Category Chart Data
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Charts")
ws5.sheet_view.showGridLines = False

ws5.row_dimensions[1].height = 10
ws5.row_dimensions[2].height = 32
ws5.merge_cells("B2:H2")
tc = ws5["B2"]
tc.value     = "Finding Distribution"
tc.fill      = fill(C["title_bg"])
tc.font      = Font(bold=True, color=C["title_fg"], size=16, name="Calibri")
tc.alignment = center()

# Severity data table
apply_header(ws5, 3, 2, "Severity")
apply_header(ws5, 3, 3, "Count")
for row_i, sev in enumerate(sev_order, start=4):
    ws5.cell(row=row_i, column=2, value=sev.upper()).fill = fill(C.get(sev, C["info"]))
    ws5.cell(row=row_i, column=2).font = font(bold=True, color="FFFFFF", size=10)
    ws5.cell(row=row_i, column=2).alignment = center()
    ws5.cell(row=row_i, column=2).border = border()
    ws5.cell(row=row_i, column=3, value=sev_counts.get(sev, 0)).fill = fill(C["alt_row"])
    ws5.cell(row=row_i, column=3).font = font(bold=True, size=11)
    ws5.cell(row=row_i, column=3).alignment = center()
    ws5.cell(row=row_i, column=3).border = border()

# Category data table
cats = {}
for r in data:
    if r["finding"]:
        cats[r["test_category"]] = cats.get(r["test_category"], 0) + 1

apply_header(ws5, 3, 5, "Category")
apply_header(ws5, 3, 6, "Findings")
for row_i, (cat, cnt) in enumerate(sorted(cats.items(), key=lambda x: -x[1]), start=4):
    ws5.cell(row=row_i, column=5, value=cat).fill = fill(C["subheader_bg"])
    ws5.cell(row=row_i, column=5).font = font(color=C["subheader_fg"], bold=True, size=9)
    ws5.cell(row=row_i, column=5).alignment = left()
    ws5.cell(row=row_i, column=5).border = border()
    ws5.cell(row=row_i, column=6, value=cnt).fill = fill(C["alt_row"])
    ws5.cell(row=row_i, column=6).font = font(bold=True, size=11)
    ws5.cell(row=row_i, column=6).alignment = center()
    ws5.cell(row=row_i, column=6).border = border()

# Bar chart — severity
bar = BarChart()
bar.type       = "col"
bar.grouping   = "clustered"
bar.title      = "Findings by Severity"
bar.y_axis.title = "Count"
bar.x_axis.title = "Severity"
bar.style      = 10
bar.width      = 14
bar.height     = 10

data_ref  = Reference(ws5, min_col=3, min_row=3, max_row=3+len(sev_order))
cats_ref  = Reference(ws5, min_col=2, min_row=4, max_row=3+len(sev_order))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.series[0].graphicalProperties.solidFill = C["critical"]

ws5.add_chart(bar, "B10")

# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
wb.save(OUT_XLSX)
print(f"\n[OK] Excel report saved: {OUT_XLSX}")
print(f"     Sheets: Executive Summary | All Test Results | Findings | Remediation Checklist | Charts")
print(f"     Records: {len(data)} tests  |  {len(findings_only)} findings")
